import io
import math
import os
import re
import tempfile
import zipfile
from dataclasses import dataclass, field

import pyreadstat
from openpyxl import load_workbook

from xml_template import assemble_xml, read_default_template

CODING_SHEET_CANDIDATES = ["קידוד", "coding"]
CATEGORY_SHEET_CANDIDATES = ["קטגוריות", "categories"]
DATAMAP_SHEET_NAME = "Datamap"
SAMPLE_SIZE = 20

BLOCK_TYPE_REGULAR = "regular"
BLOCK_TYPE_AB = "ab"
BLOCK_TYPE_CLOSED_OTHERS = "closed_others"

RAW_FILE_KIND_EXCEL = "xlsx"
RAW_FILE_KIND_SPSS = "sav"


def normalize(value):
    """Normalize a cell value for comparison: blanks -> None, numeric-looking
    values -> a canonical string (so 2, "2", 2.0 all compare equal)."""
    if value is None:
        return None
    if isinstance(value, float) and math.isnan(value):
        return None  # pandas/pyreadstat represent a blank numeric cell as NaN
    s = str(value).strip()
    if s == "":
        return None
    try:
        f = float(s)
        return str(int(f)) if f.is_integer() else str(f)
    except ValueError:
        return s


def _is_numeric_or_blank(value) -> bool:
    if value is None:
        return True
    s = str(value).strip()
    if s == "":
        return True
    try:
        float(s)
        return True
    except ValueError:
        return False


@dataclass
class QuestionBlock:
    name: str
    raw_col_index: int
    code_col_indices: list = field(default_factory=list)

    @property
    def code_count(self) -> int:
        return len(self.code_col_indices)


def _find_sheet(wb, candidates, fallback_index=0):
    for name in wb.sheetnames:
        if name.strip() in candidates:
            return wb[name]
    if fallback_index is not None and fallback_index < len(wb.sheetnames):
        return wb[wb.sheetnames[fallback_index]]
    return None


def _column_is_raw_text(col_values) -> bool:
    """A column starts a new question block if any of its first
    non-blank sampled values is not purely numeric (i.e. free text)."""
    sampled = 0
    for value in col_values:
        if value is None:
            continue
        s = str(value).strip()
        if s == "":
            continue
        sampled += 1
        if not _is_numeric_or_blank(s):
            return True
        if sampled >= SAMPLE_SIZE:
            break
    return False


def detect_blocks(ws) -> list[QuestionBlock]:
    max_row = ws.max_row
    max_col = ws.max_column
    header = [ws.cell(row=1, column=c).value for c in range(1, max_col + 1)]

    blocks: list[QuestionBlock] = []
    current: QuestionBlock | None = None

    for col in range(3, max_col + 1):
        col_values = (ws.cell(row=r, column=col).value for r in range(2, max_row + 1))
        is_raw = _column_is_raw_text(col_values) or current is None
        if is_raw:
            current = QuestionBlock(name=str(header[col - 1]), raw_col_index=col)
            blocks.append(current)
        else:
            current.code_col_indices.append(col)

    return blocks


def load_categories(wb) -> dict:
    """שאלה -> [{code, label}], read from the קטגוריות sheet."""
    ws = _find_sheet(wb, CATEGORY_SHEET_CANDIDATES, fallback_index=None)
    if ws is None:
        return {}
    categories: dict[str, list] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or len(row) < 3:
            continue
        question, code, label = row[0], row[1], row[2]
        if question is None or code is None:
            continue
        question = str(question).strip()
        categories.setdefault(question, []).append({"code": normalize(code), "label": label})
    return categories


def _union_categories(*category_lists: list[dict]) -> list[dict]:
    seen: dict = {}
    for categories in category_lists:
        for cat in categories:
            seen.setdefault(cat["code"], cat)
    return list(seen.values())


_AB_SUFFIX_RE = re.compile(r"^(.*?)([ab])$", re.IGNORECASE)

# A question's own identifier is usually a plain number ("7"), but matrix/grid
# questions get a compound, underscore-joined identifier instead (e.g. "57_60"
# for q57_60r10oe) -- so the "qN" part below is not just \d+.
_QID = r"\d+(?:_\d+)*"

# Raw Decipher data-file column naming: q{question}r{answer}oe holds the free-text
# "other" answer, q{question}r{answer} holds the closed 0/1 answers.
_OE_COL_RE = re.compile(rf"^q({_QID})r(\d+)oe$", re.IGNORECASE)
_R_COL_RE = re.compile(rf"^q({_QID})r(\d+)$", re.IGNORECASE)
# Datamap sheet header lines look like "varname: Question label text" or
# "[varname]: Question label text" (bracketed for single/text variables).
_DATAMAP_LABEL_RE = re.compile(r"^\[?(?P<name>[A-Za-z0-9_]+)\]?:\s*(?P<label>.+)$")
# OTC block name for a closed+other question is usually the bare question number
# (e.g. "q7"), but some OTC files name the block after the raw "oe" column instead
# -- either the short form ("q7oe") or the full raw-file column name ("q7r6oe").
# Either "oe" form is a strong signal the question is closed+other.
_BLOCK_QNUM_RE = re.compile(rf"^q({_QID})$", re.IGNORECASE)
_BLOCK_QNUM_OE_RE = re.compile(rf"^q({_QID})oe$", re.IGNORECASE)


def extract_block_qnum(name: str) -> tuple[str | None, bool]:
    """Returns (question_number, name_indicates_closed_others) for an OTC block
    name, matching the bare "qN", the short "qNoe", or the full "qNrMoe" (the
    raw file's own column name) naming convention."""
    m = _OE_COL_RE.match(name)
    if m:
        return m.group(1), True
    m = _BLOCK_QNUM_OE_RE.match(name)
    if m:
        return m.group(1), True
    m = _BLOCK_QNUM_RE.match(name)
    if m:
        return m.group(1), False
    return None, False


def display_question_name(name: str) -> str:
    """A trailing "oe" always marks the open-text part of a question, not a
    question in its own right -- "q7oe" is the same question as "q7", so it
    should always be labeled/exported as "q7"."""
    qnum, name_says_closed_others = extract_block_qnum(name)
    return f"q{qnum}" if name_says_closed_others else name


def suggest_ab_pairs(blocks: list[QuestionBlock]) -> dict:
    """Returns {block_name: {"base": str, "role": "A"|"B", "paired_with": str}}
    for blocks whose name ends in a/b and whose sibling (same base, opposite
    letter) also exists among the detected blocks."""
    by_name = {b.name: b for b in blocks}
    suggestions = {}
    for block in blocks:
        m = _AB_SUFFIX_RE.match(block.name)
        if not m:
            continue
        base, letter = m.group(1), m.group(2).lower()
        sibling_letter = "b" if letter == "a" else "a"
        sibling_name = base + sibling_letter
        sibling = by_name.get(sibling_name) or by_name.get(sibling_name.upper())
        if sibling is None:
            for candidate_name in by_name:
                cm = _AB_SUFFIX_RE.match(candidate_name)
                if cm and cm.group(1) == base and cm.group(2).lower() == sibling_letter:
                    sibling = by_name[candidate_name]
                    sibling_name = candidate_name
                    break
        if sibling is not None:
            suggestions[block.name] = {
                "base": base,
                "role": "A" if letter == "a" else "B",
                "paired_with": sibling_name,
            }
    return suggestions


def count_answered(ws, block: QuestionBlock) -> int:
    """Number of respondents with at least one non-blank code in this block."""
    max_row = ws.max_row
    count = 0
    for r in range(2, max_row + 1):
        if any(normalize(ws.cell(row=r, column=c).value) is not None for c in block.code_col_indices):
            count += 1
    return count


def count_answered_rows(rows: list[list]) -> int:
    """Same as count_answered, for already-built [record, uuid, *codes] rows."""
    return sum(1 for row in rows if any(normalize(v) is not None for v in row[2:]))


def build_block_rows(ws, block: QuestionBlock) -> tuple[list[str], list[list]]:
    max_row = ws.max_row
    columns = ["record", "uuid"] + [f"code{i+1}" for i in range(block.code_count)]
    rows = []
    for r in range(2, max_row + 1):
        record = ws.cell(row=r, column=1).value
        uuid = ws.cell(row=r, column=2).value
        codes = [ws.cell(row=r, column=c).value for c in block.code_col_indices]
        rows.append([record, uuid] + codes)
    return columns, rows


def rows_to_dat_bytes(columns: list[str], rows: list[list]) -> bytes:
    lines = ["\t".join(columns)]
    for row in rows:
        cells = ["" if v is None else str(v) for v in row]
        lines.append("\t".join(cells))
    return ("\r\n".join(lines) + "\r\n").encode("utf-8")


def clean_ab_pair(ws, block_a: QuestionBlock, block_b: QuestionBlock, clean_code: str):
    """Applies the unaided-awareness (TOM / others) cleaning rules to a single
    A/B pair, row by row, then re-exports the pair so that:
      - block A keeps only its (cleaned) first code column — the TOM answer.
      - block B holds every code column from both A and B, A's first (all of
        A's cleaned columns, then all of B's), representing total awareness.
    Returns ((columns_a, rows_a), (columns_b, rows_b), log_rows)."""
    max_row = ws.max_row
    clean_code = normalize(clean_code)

    def is_blank(v):
        return normalize(v) is None

    def is_clean(v):
        return normalize(v) == clean_code

    def is_real(v):
        return not is_blank(v) and not is_clean(v)

    total_code_count = block_a.code_count + block_b.code_count
    columns_a = ["record", "uuid", "code1"]
    columns_b = ["record", "uuid"] + [f"code{i+1}" for i in range(total_code_count)]
    rows_a, rows_b, log_rows = [], [], []

    for r in range(2, max_row + 1):
        record = ws.cell(row=r, column=1).value
        uid = ws.cell(row=r, column=2).value
        a = [ws.cell(row=r, column=c).value for c in block_a.code_col_indices]
        b = [ws.cell(row=r, column=c).value for c in block_b.code_col_indices]

        def log(rule, block_label, col_idx, old, new):
            log_rows.append(
                [record, uid, rule, block_label, f"code{col_idx + 1}", old, new if new is not None else ""]
            )

        # Rules 1+2: clear cleaned-answer occurrences from B when either the
        # cleaned answer appears in both A and B, or A holds a real answer.
        a_has_real = any(is_real(v) for v in a)
        a_has_clean = any(is_clean(v) for v in a)
        b_has_clean = any(is_clean(v) for v in b)
        if (a_has_clean and b_has_clean) or a_has_real:
            for i, v in enumerate(b):
                if is_clean(v):
                    log("1/2", "B", i, v, None)
                    b[i] = None

        # Rule 3: A's first col is cleaned but B's first col has a real
        # answer -> move that answer into A's first col.
        if block_a.code_count and block_b.code_count and is_clean(a[0]) and is_real(b[0]):
            log("3", "A", 0, a[0], b[0])
            log("3", "B", 0, b[0], None)
            a[0] = b[0]
            b[0] = None

        # Rule 4: within each block, a real first-col answer means any
        # cleaned answer in the other cols of that same block is erased.
        for block_label, arr in (("A", a), ("B", b)):
            if arr and is_real(arr[0]):
                for i in range(1, len(arr)):
                    if is_clean(arr[i]):
                        log("4", block_label, i, arr[i], None)
                        arr[i] = None

        # Re-export: A keeps only its (cleaned) first column; B gets every
        # column from both A and B, A's first.
        rows_a.append([record, uid, a[0] if a else None])
        rows_b.append([record, uid] + a + b)

    return (columns_a, rows_a), (columns_b, rows_b), log_rows


def log_rows_to_bytes(log_rows: list[list]) -> bytes:
    header = ["record", "uuid", "rule", "block", "column", "old_value", "new_value"]
    lines = ["\t".join(header)]
    for row in log_rows:
        cells = ["" if v is None else str(v) for v in row]
        lines.append("\t".join(cells))
    return ("\r\n".join(lines) + "\r\n").encode("utf-8")


def parse_datamap_labels(wb) -> dict[str, str]:
    """Parses the raw data file's "Datamap" sheet for question labels, keyed
    by lowercased variable name. Header lines look like "varname: label" or
    "[varname]: label" -- per-answer sub-rows (e.g. for a multi-punch
    question's r1, r2...) hold the variable name in column B instead of
    column A, so they're naturally skipped here."""
    if DATAMAP_SHEET_NAME not in wb.sheetnames:
        return {}
    ws = wb[DATAMAP_SHEET_NAME]
    labels: dict[str, str] = {}
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0] or not isinstance(row[0], str):
            continue
        m = _DATAMAP_LABEL_RE.match(row[0].strip())
        if not m:
            continue
        labels.setdefault(m.group("name").lower(), m.group("label").strip())
    return labels


def _load_excel_raw_table(file_bytes: bytes):
    """Returns (header, rows_iterator, labels) for an Excel raw data export."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
    labels = parse_datamap_labels(wb)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    header = list(next(rows_iter))
    return header, rows_iter, labels


def _pairwise_common_suffix(a: str, b: str) -> str:
    return os.path.commonprefix([a[::-1], b[::-1]])[::-1]


def _derive_multipunch_label(texts: list[str]) -> str:
    """A multi-punch question has no SPSS variable of its own -- only its
    q{id}r{M} answer columns, each labeled "{answer} - {question}". Every
    sibling shares the same question suffix, found here via the longest
    pairwise common suffix (robust to one label being cut short by SPSS's
    label-length limit) then trimmed back to just after the last " - ",
    since Hebrew's common grammatical endings can make two *different*
    answers coincidentally share a few extra trailing characters too."""
    if not texts:
        return ""
    if len(texts) == 1:
        return texts[0]
    best = ""
    for i in range(len(texts)):
        for j in range(i + 1, len(texts)):
            suffix = _pairwise_common_suffix(texts[i], texts[j])
            if len(suffix) > len(best):
                best = suffix
    idx = best.rfind(" - ")
    return (best[idx + 3 :] if idx != -1 else best).strip()


def _spss_labels(meta, header: list[str]) -> dict[str, str]:
    """Question labels from an SPSS file's own variable metadata, instead of
    a separate "Datamap" sheet. A plain single-response/text variable has a
    clean "varname: question" label. A multi-punch question's label is
    derived from its q{id}r{M} answer columns instead (see
    _derive_multipunch_label) since it has no variable of its own."""
    labels: dict[str, str] = {}
    for name, label in (meta.column_names_to_labels or {}).items():
        if not label:
            continue
        prefix = f"{name}: "
        text = label[len(prefix) :] if label.startswith(prefix) else label
        labels[name.lower()] = text.strip()

    by_qid: dict[str, list[str]] = {}
    for name in header:
        m = _R_COL_RE.match(name)
        if m:
            by_qid.setdefault(m.group(1), []).append(name)
    for qid, cols in by_qid.items():
        qname = f"q{qid}"
        if qname in labels:
            continue  # a real "q7" variable already gave us a clean label
        texts = [labels[c.lower()] for c in cols if not c.lower().endswith("oe") and c.lower() in labels]
        derived = _derive_multipunch_label(texts)
        if derived:
            labels[qname] = derived
    return labels


def _load_spss_raw_table(file_bytes: bytes):
    """Returns (header, rows_iterator, labels) for an SPSS .sav raw data
    export. pyreadstat needs a real file path rather than bytes, hence the
    temp file."""
    with tempfile.NamedTemporaryFile(suffix=".sav", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        df, meta = pyreadstat.read_sav(tmp_path)
    finally:
        os.unlink(tmp_path)
    header = list(meta.column_names)
    labels = _spss_labels(meta, header)
    rows_iter = iter(df.itertuples(index=False, name=None))
    return header, rows_iter, labels


def parse_raw_data_file(file_bytes: bytes, file_kind: str = RAW_FILE_KIND_EXCEL) -> dict:
    """Parses the raw Decipher data export (Excel or SPSS .sav). Identifies
    "closed+other" questions via q{N}r{M}oe columns, and collects the
    surviving "closed answer" columns for each such question (0/blank ->
    blank, 1 -> the answer number M from the column's own name -- e.g. a "1"
    in q7r3 becomes 3 -- the r{M} column that matches the oe column's answer
    number removed). Also resolves question labels (Excel's "Datamap" sheet,
    or the SPSS file's own variable metadata). Returns:
      {"questions": {qnum: [closed column names, in file order]},
       "rows_by_uuid": {uuid: {col_name: value}},
       "rows_by_record": {record: {col_name: value}},
       "labels": {variable_name_lowercased: label}}
    """
    if file_kind == RAW_FILE_KIND_SPSS:
        header, rows_iter, labels = _load_spss_raw_table(file_bytes)
    else:
        header, rows_iter, labels = _load_excel_raw_table(file_bytes)
    header_index = {str(h).strip(): i for i, h in enumerate(header) if h is not None}

    record_idx = header_index.get("record")
    uuid_idx = header_index.get("uuid")
    if record_idx is None and uuid_idx is None:
        raise ValueError("קובץ הנתונים חייב לכלול עמודת 'record' או 'uuid'")

    oe_answer_nums_by_question: dict[str, set] = {}
    for name in header_index:
        m = _OE_COL_RE.match(name)
        if m:
            oe_answer_nums_by_question.setdefault(m.group(1), set()).add(m.group(2))

    questions: dict[str, list[str]] = {}
    col_to_rnum: dict[str, int] = {}
    for name in header_index:
        m = _R_COL_RE.match(name)
        if not m:
            continue
        qnum, rnum = m.group(1), m.group(2)
        if qnum not in oe_answer_nums_by_question:
            continue
        if rnum in oe_answer_nums_by_question[qnum]:
            continue
        questions.setdefault(qnum, []).append(name)
        col_to_rnum[name] = int(rnum)

    for qnum, cols in questions.items():
        cols.sort(key=lambda n: header_index[n])

    keep_columns = sorted({c for cols in questions.values() for c in cols}, key=lambda n: header_index[n])
    keep_indices = [header_index[c] for c in keep_columns]

    rows_by_uuid: dict = {}
    rows_by_record: dict = {}
    all_rows: list = []
    for row in rows_iter:
        record_raw = row[record_idx] if record_idx is not None else None
        uid_raw = row[uuid_idx] if uuid_idx is not None else None
        record = normalize(record_raw)
        uid = normalize(uid_raw)
        values = {}
        for col, idx in zip(keep_columns, keep_indices):
            v = row[idx] if idx < len(row) else None
            values[col] = None if normalize(v) in (None, "0") else col_to_rnum[col]
        if uid is not None:
            rows_by_uuid[uid] = values
        if record is not None:
            rows_by_record[record] = values
        # Use the normalized record/uuid (not the raw values) in the output-facing
        # row too -- SPSS returns every numeric column as a float (e.g. 2.0), which
        # would otherwise leak a literal ".0" into the exported .dat file's columns.
        all_rows.append((record if record is not None else record_raw, uid if uid is not None else uid_raw, values))

    return {
        "questions": questions,
        "rows_by_uuid": rows_by_uuid,
        "rows_by_record": rows_by_record,
        "all_rows": all_rows,
        "labels": labels,
    }


def _otc_block_lookup(ws, block: QuestionBlock) -> tuple[dict, dict]:
    """Indexes an OTC block's coded columns by respondent, for joining onto
    another row set (the raw data file's full respondent list)."""
    by_uuid: dict = {}
    by_record: dict = {}
    for r in range(2, ws.max_row + 1):
        record = normalize(ws.cell(row=r, column=1).value)
        uid = normalize(ws.cell(row=r, column=2).value)
        codes = [ws.cell(row=r, column=c).value for c in block.code_col_indices]
        if uid is not None:
            by_uuid[uid] = codes
        if record is not None:
            by_record[record] = codes
    return by_uuid, by_record


def detect_and_describe(
    file_bytes: bytes, raw_file_bytes: bytes | None = None, raw_file_kind: str = RAW_FILE_KIND_EXCEL
) -> dict:
    """Step 1: parse the workbook and return blocks with suggested types and
    their category dictionaries, for the mapping-confirmation screen."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = _find_sheet(wb, CODING_SHEET_CANDIDATES)
    blocks = [b for b in detect_blocks(ws) if b.code_count > 0]
    categories = load_categories(wb)
    ab_suggestions = suggest_ab_pairs(blocks)
    raw_data = parse_raw_data_file(raw_file_bytes, raw_file_kind) if raw_file_bytes else None

    results = []
    for block in blocks:
        suggestion = ab_suggestions.get(block.name)
        qnum, name_says_closed_others = extract_block_qnum(block.name)
        closed_others_available = bool(raw_data is not None and qnum and qnum in raw_data["questions"])
        if suggestion:
            suggested_type = BLOCK_TYPE_AB
        elif closed_others_available or name_says_closed_others:
            suggested_type = BLOCK_TYPE_CLOSED_OTHERS
        else:
            suggested_type = BLOCK_TYPE_REGULAR
        results.append(
            {
                "name": block.name,
                "display_name": f"q{qnum}" if name_says_closed_others else block.name,
                "code_count": block.code_count,
                "answered_count": count_answered(ws, block),
                "suggested_type": suggested_type,
                "paired_with": suggestion["paired_with"] if suggestion else None,
                "role": suggestion["role"] if suggestion else None,
                "categories": categories.get(block.name, []),
                "closed_others_available": closed_others_available,
            }
        )

    return {"blocks": results}


def generate_outputs(
    file_bytes: bytes,
    mapping: list[dict],
    template_text: str | None = None,
    raw_file_bytes: bytes | None = None,
    raw_file_kind: str = RAW_FILE_KIND_EXCEL,
) -> dict:
    """Step 2: apply the confirmed type mapping and produce the .dat / log /
    XML files. `mapping` is a list of {name, type, cleaned_code?}."""
    wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    ws = _find_sheet(wb, CODING_SHEET_CANDIDATES)
    blocks = {b.name: b for b in detect_blocks(ws) if b.code_count > 0}
    mapping_by_name = {m["name"]: m for m in mapping if m["name"] in blocks}
    categories = load_categories(wb)
    raw_data = parse_raw_data_file(raw_file_bytes, raw_file_kind) if raw_file_bytes else None

    results = []
    dat_files = {}
    warnings = []
    handled = set()
    xml_question_entries = []

    labels = raw_data["labels"] if raw_data else {}

    def question_label(name: str) -> str | None:
        return labels.get(name.lower())

    ab_entries = [m for m in mapping_by_name.values() if m.get("type") == BLOCK_TYPE_AB]
    ab_suggestions = suggest_ab_pairs(list(blocks.values()))

    for entry in ab_entries:
        name = entry["name"]
        if name in handled:
            continue
        suggestion = ab_suggestions.get(name)
        if not suggestion:
            warnings.append(f"{name}: לא נמצא זוג AB תואם (שם עם סיומת a/b) — יטופל כשאלה רגילה")
            continue
        paired_name = suggestion["paired_with"]
        paired_entry = mapping_by_name.get(paired_name)
        if not paired_entry or paired_entry.get("type") != BLOCK_TYPE_AB:
            warnings.append(f"{name}: השאלה המזווגת {paired_name} לא סומנה כ-AB — יטופל כשאלה רגילה")
            continue

        clean_code = entry.get("cleaned_code") or paired_entry.get("cleaned_code")
        if not clean_code:
            warnings.append(f"{name}/{paired_name}: לא נבחרה 'תשובה לניקוי' — יטופל כשאלה רגילה")
            continue

        block_a = blocks[name] if suggestion["role"] == "A" else blocks[paired_name]
        block_b = blocks[paired_name] if suggestion["role"] == "A" else blocks[name]
        (cols_a, rows_a), (cols_b, rows_b), log_rows = clean_ab_pair(ws, block_a, block_b, clean_code)

        base = suggestion["base"]
        fname_a = f"{block_a.name}_coded.dat"
        fname_b = f"{block_b.name}_coded.dat"
        fname_log = f"{base}_cleaning_log.txt"
        dat_files[fname_a] = rows_to_dat_bytes(cols_a, rows_a)
        dat_files[fname_b] = rows_to_dat_bytes(cols_b, rows_b)
        dat_files[fname_log] = log_rows_to_bytes(log_rows)

        categories_a = categories.get(block_a.name, [])
        categories_b = categories.get(block_b.name, [])
        categories_union = _union_categories(categories_a, categories_b)

        results.append(
            {
                "question_name": block_a.name,
                "type": BLOCK_TYPE_AB,
                "role": "A",
                "code_count": len(cols_a) - 2,
                "category_count": len(categories_a),
                "row_count": len(rows_a),
                "answered_count": count_answered_rows(rows_a),
                "filename": fname_a,
                "columns": cols_a,
                "preview_rows": rows_a[:20],
            }
        )
        results.append(
            {
                "question_name": block_b.name,
                "type": BLOCK_TYPE_AB,
                "role": "B",
                "code_count": len(cols_b) - 2,
                "category_count": len(categories_union),
                "row_count": len(rows_b),
                "answered_count": count_answered_rows(rows_b),
                "filename": fname_b,
                "columns": cols_b,
                "preview_rows": rows_b[:20],
            }
        )
        results.append(
            {
                "question_name": base,
                "type": "log",
                "code_count": None,
                "category_count": None,
                "row_count": len(log_rows),
                "answered_count": None,
                "filename": fname_log,
                "columns": ["record", "uuid", "rule", "block", "column", "old_value", "new_value"],
                "preview_rows": log_rows[:20],
            }
        )
        handled.add(name)
        handled.add(paired_name)
        xml_question_entries.append(
            {
                "name": f"{block_a.name}_coded",
                "code_count": len(cols_a) - 2,
                "categories": categories_a,
                "label": question_label(block_a.name),
            }
        )
        xml_question_entries.append(
            {
                "name": f"{block_b.name}_coded",
                "code_count": len(cols_b) - 2,
                "categories": categories_union,
                "label": question_label(block_b.name),
            }
        )

    closed_others_entries = [m for m in mapping_by_name.values() if m.get("type") == BLOCK_TYPE_CLOSED_OTHERS]

    # Group by resolved display name first -- more than one OTC block can
    # belong to the same underlying question (e.g. two separate "other"
    # slots, q15r6oe and q15r9oe, both belong to question 15) and must be
    # merged into a single export, not each produce a competing duplicate
    # "q15_coded" file/XML block.
    closed_others_groups: dict[str, list[dict]] = {}
    for entry in closed_others_entries:
        name = entry["name"]
        if name in handled or name not in blocks:
            continue
        qnum, _ = extract_block_qnum(name)
        display_name = display_question_name(name)
        closed_others_groups.setdefault(display_name, []).append(
            {"name": name, "qnum": qnum, "block": blocks[name]}
        )

    for display_name, group_entries in closed_others_groups.items():
        qnum = group_entries[0]["qnum"]

        if raw_data is None:
            warnings.append(f"{display_name}: לא נטען קובץ נתונים גולמי — השאלה תטופל כשאלה רגילה")
            continue
        if not qnum:
            warnings.append(
                f"{display_name}: שם השאלה אינו בפורמט qN או qNoe, לא ניתן להתאים לקובץ הנתונים — תטופל כשאלה רגילה"
            )
            continue
        if qnum not in raw_data["questions"]:
            warnings.append(
                f"{display_name}: לא נמצאה קבוצת עמודות 'סגורה+אחר' תואמת (q{qnum}r..oe) בקובץ הנתונים — תטופל כשאלה רגילה"
            )
            continue

        closed_columns = raw_data["questions"].get(qnum, [])
        otc_lookups = [(_otc_block_lookup(ws, e["block"]), e["block"]) for e in group_entries]
        otc_code_count = sum(e["block"].code_count for e in group_entries)
        total_code_count = len(closed_columns) + otc_code_count
        columns = ["record", "uuid"] + [f"code{i+1}" for i in range(total_code_count)]

        rows = []
        for record, uid, values in raw_data["all_rows"]:
            closed_values = [values.get(c) for c in closed_columns]
            all_codes = []
            for (otc_by_uuid, otc_by_record), block in otc_lookups:
                codes = otc_by_uuid.get(normalize(uid))
                if codes is None:
                    codes = otc_by_record.get(normalize(record))
                if codes is None:
                    codes = [None] * block.code_count
                all_codes.extend(codes)
            rows.append([record, uid] + closed_values + all_codes)

        filename = f"{display_name}_coded.dat"
        dat_files[filename] = rows_to_dat_bytes(columns, rows)
        combined_categories = _union_categories(*(categories.get(e["block"].name, []) for e in group_entries))
        results.append(
            {
                "question_name": display_name,
                "type": BLOCK_TYPE_CLOSED_OTHERS,
                "role": None,
                "code_count": len(columns) - 2,
                "category_count": len(combined_categories),
                "row_count": len(rows),
                "answered_count": count_answered_rows(rows),
                "filename": filename,
                "columns": columns,
                "preview_rows": rows[:20],
            }
        )
        for e in group_entries:
            handled.add(e["name"])
        xml_question_entries.append(
            {
                "name": f"{display_name}_coded",
                "code_count": len(columns) - 2,
                "categories": combined_categories,
                "label": question_label(display_name),
            }
        )

    for name, block in blocks.items():
        if name in handled:
            continue
        display_name = display_question_name(name)
        columns, rows = build_block_rows(ws, block)
        filename = f"{display_name}_coded.dat"
        dat_files[filename] = rows_to_dat_bytes(columns, rows)
        entry = mapping_by_name.get(name, {})
        results.append(
            {
                "question_name": display_name,
                "type": entry.get("type", BLOCK_TYPE_REGULAR),
                "role": None,
                "code_count": block.code_count,
                "category_count": len(categories.get(block.name, [])),
                "row_count": len(rows),
                "answered_count": count_answered_rows(rows),
                "filename": filename,
                "columns": columns,
                "preview_rows": rows[:20],
            }
        )
        xml_question_entries.append(
            {
                "name": f"{display_name}_coded",
                "code_count": block.code_count,
                "categories": categories.get(block.name, []),
                "label": question_label(display_name),
            }
        )

    if template_text is None:
        template_text = read_default_template()
    xml_bytes, xml_warnings = assemble_xml(template_text, xml_question_entries)
    warnings.extend(xml_warnings)
    if xml_bytes is not None:
        xml_filename = "survey_openends.xml"
        dat_files[xml_filename] = xml_bytes
        results.append(
            {
                "question_name": "XML",
                "type": "xml",
                "role": None,
                "code_count": None,
                "category_count": None,
                "row_count": None,
                "answered_count": None,
                "filename": xml_filename,
                "columns": None,
                "preview_rows": None,
                "text_preview": xml_bytes.decode("utf-8")[:4000],
            }
        )

    return {"blocks": results, "dat_files": dat_files, "warnings": warnings}


def zip_dat_files(dat_files: dict) -> bytes:
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for filename, data in dat_files.items():
            zf.writestr(filename, data)
    return buf.getvalue()
